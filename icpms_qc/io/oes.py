"""Agilent ICP Expert workbook (ICP-OES) → Batch.

Why a third reader
------------------
Optical emission is close enough to mass spectrometry that every QC criterion in
the catalog still applies, and far enough that neither existing reader can touch
it. Three differences, in increasing order of how much they matter:

* It is a **workbook**, not a CSV, and each measure is its own sheet rather than
  a column suffix: `Conc. in Sample Units`, `Corrected Intensities`,
  `... RSDs`, `Internal Standards`. They are read and merged back into one batch.

* An analyte is identified by its **emission wavelength**, not by a mass:
  `U 385.958` rather than `238 U [No Gas]`. There is no mass number to parse
  because the technique has no such concept.

* The same element is routinely measured on **several lines at once**, and their
  results are expected to agree. That is a whole QC question mass spectrometry
  does not have, and it is what `oes_line_agreement` exists for.

Sample types are not stated. The export carries a `QC Status` column but no type,
so the template's name patterns decide, exactly as for the Element, and anything
they do not claim stays OTHER with a warning.
"""
from __future__ import annotations

import re

from icpms_qc.io import templates
from icpms_qc.model import Analyte, Batch, Result, Sample, SampleType

#: "U 385.958", "Fe 259.940 (cps)", "U 385.958\n(cps)" — element then wavelength.
LINE_RE = re.compile(r"^\s*(?P<element>[A-Z][a-z]?)\s+(?P<nm>\d{2,3}\.\d{1,4})"
                     r"(?:\s*\(\s*(?P<unit>[^)]+?)\s*\))?")

SHEET_CONC = "Conc. in Sample Units"
SHEET_CONC_ALT = "Conc. in Calib Units"
SHEET_INTENSITY = "Corrected Intensities"
SHEET_RSD = "Corrected Intensities RSDs"
SHEET_ISTD = "Internal Standards"

#: Columns before the analytes begin.
_INFO = {"sample id", "r", "acquisition time", "qc status",
         "dataset file", "method file", "solution type", "rack:tube"}

_WS = re.compile(r"\s+")


def analyte_from_line(text: str) -> Analyte | None:
    """`U 385.958 (cps)` → Analyte(element=U, wavelength_nm=385.958)."""
    flat = _WS.sub(" ", str(text or "")).strip()
    m = LINE_RE.match(flat)
    if not m:
        return None
    label = f"{m.group('element')} {m.group('nm')}"
    return Analyte(label=label, element=m.group("element"),
                   wavelength_nm=float(m.group("nm")))


def unit_from_line(text: str) -> str | None:
    """The unit rides in the header: `U 385.958 (mg/L)`. Read it, do not assume."""
    m = LINE_RE.match(_WS.sub(" ", str(text or "")).strip())
    u = m.group("unit") if m else None
    return None if not u or u.lower() == "cps" else u


def _num(v) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in {"-", "N/A", "n/a", "NA"}:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _sheet(wb, name: str):
    for actual in wb.sheetnames:
        if actual.strip().lower() == name.strip().lower():
            return wb[actual]
    return None


def _read_sheet(ws) -> tuple[list[str], list[list]]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return [], []
    header = [_WS.sub(" ", "" if c is None else str(c)).strip() for c in rows[0]]
    return header, rows[1:]


def looks_like_oes_workbook(path: str) -> bool:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    names = {n.strip().lower() for n in wb.sheetnames}
    return any(s.lower() in names for s in (SHEET_CONC, SHEET_CONC_ALT, SHEET_INTENSITY))


def parse(path: str, template: str | None = "agilent_oes") -> Batch:
    """Read an ICP Expert workbook, merging its measure sheets into one batch."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:                       # pragma: no cover
        raise ValueError(
            "reading ICP-OES workbooks needs openpyxl: pip install 'icpms-qc[oes]'"
        ) from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tpl = templates.load(template) if template else None

    primary = _sheet(wb, SHEET_CONC) or _sheet(wb, SHEET_CONC_ALT) \
        or _sheet(wb, SHEET_INTENSITY)
    if primary is None:
        raise ValueError(
            f"{path}: not an ICP Expert workbook — expected a "
            f"'{SHEET_CONC}' or '{SHEET_INTENSITY}' sheet, found "
            f"{', '.join(wb.sheetnames[:6])}")

    header, body = _read_sheet(primary)
    if not body:
        raise ValueError(f"{path}: sheet '{primary.title}' has no data rows")

    name_i = next((i for i, h in enumerate(header)
                   if h.strip().lower() == "sample id"), None)
    if name_i is None:
        raise ValueError(f"{path}: no 'Sample Id' column in '{primary.title}'")

    analytes: dict[int, Analyte] = {}
    units: dict[int, str] = {}
    for i, h in enumerate(header):
        if i == name_i or h.strip().lower() in _INFO:
            continue
        if (a := analyte_from_line(h)) is not None:
            analytes[i] = a
            if (u := unit_from_line(h)):
                units[i] = u
    if not analytes:
        raise ValueError(f"{path}: no emission-line columns in '{primary.title}'")

    batch = Batch(source_path=str(path), template_id=tpl.id if tpl else "agilent_oes",
                  instrument_family="agilent-icp-oes")
    batch.analytes = list(analytes.values())

    # The other sheets repeat the same grid, so they are indexed by row position
    # and by the same column offsets rather than re-derived.
    extra = {}
    for key, sheet in (("intensity", SHEET_INTENSITY), ("rsd", SHEET_RSD)):
        ws = _sheet(wb, sheet)
        if ws is None or ws is primary:
            continue
        h2, b2 = _read_sheet(ws)
        # Headers differ between sheets only by the unit they carry, "(cps)"
        # against "(mg/L)", so they are matched on the lines themselves.
        if [analyte_from_line(h) for h in h2] == [analyte_from_line(h) for h in header]:
            extra[key] = b2
        else:
            batch.warnings.append(
                f"sheet '{sheet}' lists different emission lines and was skipped")

    unknown: set[str] = set()
    for r, row in enumerate(body):
        name = "" if r >= len(body) or name_i >= len(row) or row[name_i] is None \
            else str(row[name_i]).strip()
        if not name and not any(_num(row[i]) is not None for i in analytes if i < len(row)):
            continue                                   # padding row at the foot

        stype = tpl.sample_type_vocab.get(name) if tpl else None
        if stype is None and tpl:
            for pat, t in tpl.sample_type_patterns:
                if pat.search(name):
                    stype = t
                    break
        if stype is None and tpl and tpl.default_sample_type:
            stype = tpl.default_sample_type
        if stype is None:
            stype = SampleType.OTHER
            if name not in unknown:
                unknown.add(name)
                batch.warnings.append(
                    f"unrecognized sample name '{name}' -> OTHER "
                    f"(add it to the template's sample_type_patterns)")

        sample = Sample(name=name or f"row{r + 1}", seq_index=r + 1, type=stype)
        if tpl and tpl.expected_conc_from_name and stype in {
                SampleType.CAL_STD, SampleType.ICV, SampleType.CCV, SampleType.LCS}:
            if m := tpl.expected_conc_from_name.search(name):
                sample.level = float(m.group("value"))

        for i, a in analytes.items():
            def other(key: str):
                rows = extra.get(key)
                if rows is None or r >= len(rows) or i >= len(rows[r]):
                    return None
                return _num(rows[r][i])
            sample.results[a.label] = Result(
                conc=_num(row[i]) if i < len(row) else None,
                unit=units.get(i) or (str(tpl.columns.get("unit", "ppb")) if tpl else "ppb"),
                intensity=other("intensity"),
                rsd_pct=other("rsd"),
            )
        batch.samples.append(sample)

    return batch
